import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


def write_df_with_metadata(ws, df, meta_dict):
    # Write metadata rows (max 10 rows)
    for i, (k, v) in enumerate(meta_dict.items()):
        ws.cell(row=i + 1, column=1, value=k)
        ws.cell(row=i + 1, column=2, value=str(v))
        ws.cell(row=i + 1, column=1).alignment = Alignment(horizontal="right")

    # Write data starting at row 11
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=11):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)



def ds_to_excel(ds,excel_path):
    # Prepare the Excel file
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group 1D variables
    one_d_vars = []
    scalars = {}

    for var in ds.data_vars:
        da = ds[var]

        meta = {
            "Variable": var,
            "Dimensions": ", ".join(da.dims),
            "Shape": da.shape,
            "Units": da.attrs.get("units", ""),
            "Long Name": da.attrs.get("long_name", ""),
            "Standard Name": da.attrs.get("standard_name", ""),
            "Other Attributes": {k: v for k, v in da.attrs.items() if k not in ["units", "long_name", "standard_name"]}
        }

        if da.ndim == 0:
            scalars[var] = da.values.item()
        elif da.ndim == 1:
            one_d_vars.append(var)
        else:
            # Multi-D variable → new sheet
            df = da.to_dataframe().reset_index()
            if len(df) > 10000:
                #if the file is too large, will only keep top 10k rows
                df = df.iloc[:10000]
            ws = wb.create_sheet(title=var[:31])
            write_df_with_metadata(ws, df, meta)

    # Combine 1D variables into one sheet with metadata
    if one_d_vars:
        df_1d = ds[one_d_vars].to_dataframe().reset_index()
        meta = {
            "Sheet": "1D_Variables",
            "Variables Included": ", ".join(one_d_vars)
        }
        ws_1d = wb.create_sheet("1D_Variables")
        write_df_with_metadata(ws_1d, df_1d, meta)

    # Add scalar variables to a sheet
    if scalars:
        df_scalar = pd.DataFrame(list(scalars.items()), columns=["Variable", "Value"])
        ws_scalars = wb.create_sheet("Scalars")
        for r_idx, row in enumerate(dataframe_to_rows(df_scalar, index=False, header=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                ws_scalars.cell(row=r_idx, column=c_idx, value=val)

    # Save final Excel file
    wb.save(excel_path)
    print(f"Excel saved with embedded metadata at: {excel_path}")